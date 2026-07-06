# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "00277F"
GOLD = "F0C157"
WHITE = "FFFFFF"
SLATE = "5E5E5E"
LIGHTBG = "F4F6FB"

PHASE_COLORS = {
    "1 · Spark":            "FFE8C2",
    "2 · Practice":         "D9E6FF",
    "3 · Ideas":            "D8F0DC",
    "4 · Build-Readiness":  "F3D9F5",
    "5 · Countdown":        "FFD3CE",
}

rows = [
# n, date, phase, title, format/angle, skill or asset built, excitement mechanic
(1, "Thu Jul 09, 2026", "1 · Spark", "Meet Your First AI Teammate",
 "5-minute walkthrough of the one approved tool, framed as introducing a new colleague, not a training manual.",
 "First hands-on exposure, zero prior skill assumed.",
 "Removes first-use fear before it forms; nothing to fail at yet."),
(2, "Thu Jul 16, 2026", "1 · Spark", "The Five-Minute Ceiling",
 "Live before/after on one real Finance task, done in exactly five minutes on camera or in a short clip.",
 "Launches the Prompt Library (entry #1).",
 "Kills the '‘I don’t have time’' objection with proof, not persuasion."),
(3, "Thu Jul 23, 2026", "1 · Spark", "Who’s Already Doing This?",
 "Spotlight 1–2 real Finance colleagues already experimenting informally — named, quoted, specific.",
 "Surfaces existing grassroots users as an informal talent pool.",
 "Internal social proof: ‘people like me’ beats any outside case study."),
(4, "Thu Jul 30, 2026", "1 · Spark", "Announcing: Vanguard Hack, December 2026",
 "Formal reveal as an invitation, not a mandate. Opens a standing ‘submit a problem you’d want solved’ form.",
 "Opens the problem-submission channel (stays open for months).",
 "Starts the countdown clock; ownership via open call instead of top-down assignment."),
(5, "Thu Aug 06, 2026", "1 · Spark", "What a Hackathon Actually Is (and Isn’t)",
 "Demystifies the format for first-timers: no all-nighters, mixed-skill teams, trying beats winning.",
 "Shared mental model of what December will actually look like.",
 "Neutralizes intimidation before it turns into quiet opt-out."),
(6, "Thu Aug 13, 2026", "1 · Spark", "The Governance Gate — and Why It Protects You",
 "Reframes the intake/approval process as what makes it safe to experiment, not red tape.",
 "Publishes the real intake form and the 3-person review contacts.",
 "Trust-building: ‘I won’t get in trouble for trying’ is a precondition for real excitement."),

(7, "Thu Aug 20, 2026", "2 · Practice", "Your First AI-Assisted Task, Start to Finish",
 "Full step-by-step walkthrough on one real, low-risk task readers can follow this week.",
 "Prompt Library Vol. 2.",
 "Converts spectators into doers with a task they can literally copy."),
(8, "Thu Aug 27, 2026", "2 · Practice", "The Tool Catalog, Explained",
 "One-page cheat sheet: every approved tool, what it’s actually good for, what it isn’t.",
 "Reference asset for picking the right tool without guessing.",
 "Removes decision paralysis — fewer excuses not to start."),
(9, "Thu Sep 03, 2026", "2 · Practice", "30-Minute Challenge: Automate One Annoying Thing",
 "First optional individual mini-challenge. No team required. Submit what you built.",
 "First individually-authored artifact from each participant.",
 "Gamified but low-stakes; plants the first personal ‘I made something’ moment."),
(10, "Thu Sep 10, 2026", "2 · Practice", "Challenge Results: What You Built",
 "Features real reader submissions from the 30-minute challenge, credited by name.",
 "Public recognition + peer-to-peer idea spread.",
 "Seeing named colleagues’ real work is the single strongest adoption trigger available."),
(11, "Thu Sep 17, 2026", "2 · Practice", "Fails Worth Sharing",
 "Normalizes a real attempt that didn’t work, and what was learned from it.",
 "Establishes failure as data, not embarrassment.",
 "Directly defuses the #1 quiet reason people skip hackathons: fear of looking incompetent."),
(12, "Thu Sep 24, 2026", "2 · Practice", "Prompt Library, Vol. 3: Reader-Submitted",
 "Crowdsourced additions to the library, credited by name.",
 "Library now visibly built by the community, not just Vanguard.",
 "Ownership shifts from ‘content we’re given’ to ‘content we built.’"),

(13, "Thu Oct 01, 2026", "3 · Ideas", "What’s the Most Annoying Part of Your Month-End?",
 "Short survey/prompt issue that harvests real recurring pain points directly from readers.",
 "Raw material bank for hackathon problem statements.",
 "People get excited about solving their own named frustration, not an assigned brief."),
(14, "Thu Oct 08, 2026", "3 · Ideas", "Vanguard Hack: The Problem Shortlist",
 "Publishes a curated shortlist of pain points sourced directly from reader submissions.",
 "Official hackathon problem menu, community-authored.",
 "Readers see their own words reflected back as legitimate, sanctioned problems to solve."),
(15, "Thu Oct 15, 2026", "3 · Ideas", "Find Your Team",
 "Lightweight, informal team-formation mechanic — a skill/interest form or an organized pairing mixer.",
 "Registered teams and rosters.",
 "Bottom-up team formation beats top-down assignment for genuine buy-in."),
(16, "Thu Oct 22, 2026", "3 · Ideas", "Borrowed From the Outside",
 "Introduces the Moderna / JPMorgan / McKinsey / Commonwealth Bank benchmarks — now that readers have their own experience to compare against.",
 "External proof points reframed as inspiration, not theory.",
 "Case studies land only once people have their own frame of reference to test them against."),

(17, "Thu Oct 29, 2026", "4 · Build-Readiness", "How to Scope a One-Day Build",
 "Practical guidance on picking a problem small enough to actually finish inside a hackathon window.",
 "Scoping framework / worksheet each team fills out before December.",
 "Removes the single biggest cause of hackathon-day failure: over-ambition."),
(18, "Thu Nov 05, 2026", "4 · Build-Readiness", "Pitch Like You Mean It",
 "A 3-minute pitch structure borrowed from how product teams and VCs actually pitch, not a generic slide template.",
 "Pitch template every team can reuse.",
 "Confidence in the final act removes a major source of pre-event dread."),
(19, "Thu Nov 12, 2026", "4 · Build-Readiness", "Dry Run: The Warm-Up Sprint",
 "Announces and recaps an optional half-day, low-stakes practice mini-hackathon ahead of the real one.",
 "First full rehearsal of the build-and-pitch cycle.",
 "Burns off first-timer anxiety on a low-stakes day so December feels familiar, not novel."),
(20, "Thu Nov 19, 2026", "4 · Build-Readiness", "What Judges Are Actually Looking For",
 "Transparent breakdown of the real evaluation criteria.",
 "Shared understanding of what ‘good’ looks like.",
 "Removes anxiety from guessing the rules of a game nobody explained."),

(21, "Thu Nov 26, 2026", "5 · Countdown", "Meet the Teams",
 "Spotlights every registered team and the problem they chose, one by one.",
 "Public roster and mutual visibility across the whole Finance org.",
 "Seeing the full field assembled is itself a momentum device."),
(22, "Thu Dec 03, 2026", "5 · Countdown", "Leadership Is Playing Too",
 "Features execs and sponsors mentoring, fielding their own team, or confirmed as judges.",
 "Visible top-level psychological safety signal.",
 "If leadership is seen trying and risking looking silly too, it is genuinely safe for everyone else."),
(23, "Thu Dec 10, 2026", "5 · Countdown", "T-Minus: Final Logistics",
 "Schedule, what to bring, FAQ, last reminders — practical, not motivational.",
 "Zero-friction arrival on hackathon day.",
 "Removes logistical anxiety so energy goes into building, not figuring out where to be."),
(24, "Week after Vanguard Hack", "Post-Event", "What We Built",
 "Recap issue: every team’s output, what surprised people, and which builds are worth carrying into real workflows.",
 "First candidates for adoption into the Vanguard platform — the loop back to ‘measure reuse, not adoption.’",
 "Converts a one-time event into the opening chapter of the next cycle, not a closed loop."),
]

wb = Workbook()
ws = wb.active
ws.title = "Editorial Calendar"

headers = ["#", "Date", "Phase", "Issue Title", "Format / Angle", "Skill or Asset Built", "Excitement Mechanic (the 'for real' part)"]
ws.append(headers)
for col in range(1, len(headers) + 1):
    c = ws.cell(row=1, column=col)
    c.font = Font(name="Arial", bold=True, color=WHITE, size=11)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

thin = Side(style="thin", color="D0D5E3")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for r in rows:
    ws.append(r)
    row_idx = ws.max_row
    phase = r[2]
    fill_color = PHASE_COLORS.get(phase, "FFFFFF") if phase != "Post-Event" else "E4E4E4"
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=row_idx, column=col)
        c.font = Font(name="Arial", size=10, color="1A1A1A" if col != 1 else NAVY,
                       bold=(col == 1))
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = border
        if col in (3,):
            c.fill = PatternFill("solid", fgColor=fill_color)
            c.font = Font(name="Arial", size=10, bold=True, color="1A1A1A")
        elif row_idx % 2 == 0:
            c.fill = PatternFill("solid", fgColor=LIGHTBG)

widths = {1: 5, 2: 16, 3: 16, 4: 30, 5: 46, 6: 30, 7: 40}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

for r_idx in range(2, ws.max_row + 1):
    ws.row_dimensions[r_idx].height = 60

# ---------------------------------------------------------------- Design Notes sheet
ws2 = wb.create_sheet("Design Notes")
ws2.column_dimensions["A"].width = 34
ws2.column_dimensions["B"].width = 100

def section(title, body_lines, row):
    c = ws2.cell(row=row, column=1, value=title)
    c.font = Font(name="Arial", bold=True, size=12, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c2 = ws2.cell(row=row, column=2)
    c2.fill = PatternFill("solid", fgColor=NAVY)
    row += 1
    for line in body_lines:
        cell = ws2.cell(row=row, column=1, value=line[0])
        cell.font = Font(name="Arial", bold=True, size=10, color=NAVY)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell2 = ws2.cell(row=row, column=2, value=line[1])
        cell2.font = Font(name="Arial", size=10, color="1A1A1A")
        cell2.alignment = Alignment(vertical="top", wrap_text=True)
        ws2.row_dimensions[row].height = 46
        row += 1
    return row + 1

row = 1
row = section("THE ARC (23 weekly issues, Jul 9 – Dec 10, 2026)", [
    ("1 · Spark (Wks 1–6)", "Curiosity and identity. Goal: make trying feel low-risk and normal, not assigned."),
    ("2 · Practice (Wks 7–12)", "Hands-on reps on real (small) tasks. Goal: individual, visible, named wins."),
    ("3 · Ideas (Wks 13–16)", "Shift from learning AI to naming real problems. Goal: reader-sourced hackathon shortlist and teams."),
    ("4 · Build-Readiness (Wks 17–20)", "Practical hackathon craft: scoping, pitching, a low-stakes dry run. Goal: competence removes anxiety."),
    ("5 · Countdown (Wks 21–23)", "Final momentum and logistics. Goal: visible full field, visible leadership, zero friction on the day."),
], row)

row = section("WHY THIS BUILDS REAL EXCITEMENT, NOT HYPE", [
    ("Cumulative visible artifacts", "The Prompt Library and the Problem Shortlist grow issue over issue. People can watch progress accumulate and want to add to it — proof beats promises."),
    ("Internal social proof over case studies", "Named colleagues' real attempts (issues 3, 10, 11) beat Moderna/JPMorgan every time. Outside benchmarks are saved for issue 16, after readers have their own frame of reference."),
    ("Reader-sourced content", "Problems (issue 13-14), prompts (issue 12), and teams (issue 15) are sourced from readers, not handed to them. Ownership beats being told what to do."),
    ("Explicit anxiety removal", "Issues 5, 6, 11, 17, 19, 20 exist specifically to name and defuse the quiet reasons people skip hackathons: fear of wasting time, fear of breaking rules, fear of looking incompetent, fear of not knowing what 'good' means."),
    ("Staged reveals, not one announcement", "The hackathon is announced in issue 4, given real problems in issue 14, given real teams in issue 21, and given visible leadership buy-in in issue 22 — anticipation is built in stages, mirroring how real launches create momentum."),
    ("Leadership visibly at risk too", "Issue 22 only works if it's true. If no leader is willing to be seen trying and possibly failing, do not run this issue — find one who will, first."),
], row)

row = section("WHAT NOT TO DO", [
    ("Don't skip Phase 1/2 to save time", "Jumping straight to 'submit your hackathon idea' without 6+ weeks of low-stakes practice first produces a hackathon of the same 5 people who were already comfortable with AI."),
    ("Don't make the 30-min challenge or dry run mandatory", "The moment participation becomes compliance, the excitement mechanic inverts. Optional-but-visible is the entire design."),
    ("Don't let 'Borrowed From the Outside' run first", "Case studies land as inspiration only after readers have their own experience. Run early, they read as corporate messaging."),
], row)

wb.save("Finance_Forward_Hackathon_Catalogue.xlsx")
print("saved", ws.max_row - 1, "issues")
